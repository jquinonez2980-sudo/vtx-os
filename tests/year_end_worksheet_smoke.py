"""
tests/year_end_worksheet_smoke.py
Offline smoke tests for core/year_end_worksheet.py.

Requires the real template at:
    R:\\Templates\\TEMPLATE_YearEnd_Accounting_Professional_BLANK_v2.xlsx

Creates a temp output in the system temp dir; no R: drive writes during normal run.

Run:
    python tests/year_end_worksheet_smoke.py
"""

from __future__ import annotations

import sys
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from core.year_end_worksheet import populate_worksheet
from sage50.trial_balance_parser import TBLine

TEMPLATE_PATH = Path(r"R:\Templates\TEMPLATE_YearEnd_Accounting_Professional_BLANK_v2.xlsx")

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SYNTHETIC_TB: list[TBLine] = [
    TBLine("1060", "Bank - TD Chequing",        Decimal("12202.87"), Decimal("0")),
    TBLine("1200", "Accounts Receivable",        Decimal("5000.00"),  Decimal("0")),
    TBLine("2000", "Accounts Payable",           Decimal("0"),        Decimal("3800.00")),
    TBLine("3100", "Retained Earnings",          Decimal("0"),        Decimal("8402.87")),
    TBLine("4100", "Revenue",                    Decimal("0"),        Decimal("75000.00")),
    TBLine("5200", "Bank Charges & Interest",    Decimal("228.00"),   Decimal("0")),
    TBLine("5400", "Insurance",                  Decimal("2184.00"),  Decimal("0")),
    TBLine("5600", "Telephone & Cellular",       Decimal("1107.00"),  Decimal("0")),
    TBLine("5850", "Wages & Benefits",           Decimal("66481.00"), Decimal("0")),
]

CLIENT_NAME   = "Concetta Enterprises Inc."
YEAR_END_DATE = "April 30, 2026"
PREPARED_BY   = "Jorge Quinonez CPA"

# ---------------------------------------------------------------------------
# Test runner
# ---------------------------------------------------------------------------

passed = 0
failed = 0


def check(label: str, condition: bool) -> None:
    global passed, failed
    if condition:
        print(f"  PASS  {label}")
        passed += 1
    else:
        print(f"  FAIL  {label}")
        failed += 1


print("=== year_end_worksheet_smoke ===\n")

# --- Template availability ---
print("0. Template availability")
if not TEMPLATE_PATH.exists():
    print(f"  SKIP  Template not found at {TEMPLATE_PATH} — cannot run tests")
    print("        Export the template from R:\\Templates before running this test.")
    sys.exit(0)
check("template file exists", TEMPLATE_PATH.exists())

# --- Generate output ---
print("\n1. populate_worksheet runs without error")
with tempfile.TemporaryDirectory() as tmp:
    out = Path(tmp) / "concetta_yearend_2026-04.xlsx"
    try:
        result = populate_worksheet(
            template_path=TEMPLATE_PATH,
            output_path=out,
            client_name=CLIENT_NAME,
            year_end_date=YEAR_END_DATE,
            tb_lines=SYNTHETIC_TB,
            prepared_by=PREPARED_BY,
        )
        check("returns output_path", result == out)
        check("output file exists", out.exists())

        # --- Verify Cover Sheet ---
        print("\n2. Cover Sheet values")
        wb = openpyxl.load_workbook(out, data_only=True)
        cover = wb["0. Cover Sheet"]
        check("D8 = client name",    cover["D8"].value == CLIENT_NAME)
        check("D9 = year end date",  cover["D9"].value == YEAR_END_DATE)
        check("D10 = prepared by",   cover["D10"].value == PREPARED_BY)
        check("D11 = date (set)",    cover["D11"].value is not None)

        # --- Verify Worksheet tab ---
        print("\n3. Worksheet tab TB data")
        ws = wb["1. Worksheet"]
        check("A2 = first account no",  ws["A2"].value == "1060")
        check("B2 = first description", ws["B2"].value == "Bank - TD Chequing")
        check("C2 = first debit",       abs(ws["C2"].value - 12202.87) < 0.005)
        check("D2 is None (no credit)", ws["D2"].value is None)

        # Second row — Accounts Receivable
        check("A3 = 1200",   ws["A3"].value == "1200")
        check("C3 = 5000",   abs(ws["C3"].value - 5000.0) < 0.005)

        # A liability row — credit only
        check("A4 = 2000 (AP)",          ws["A4"].value == "2000")
        check("C4 is None (no debit)",   ws["C4"].value is None)
        check("D4 = 3800 (credit)",      abs(ws["D4"].value - 3800.0) < 0.005)

        # Row count — 9 TB lines should occupy rows 2-10
        last_row = 2 + len(SYNTHETIC_TB) - 1
        check(f"last row A{last_row} populated",
              ws.cell(row=last_row, column=1).value is not None)
        # Row after last should be empty in cols A-D
        check(f"row {last_row+1} col A is empty",
              ws.cell(row=last_row + 1, column=1).value is None)

        # E column must NOT be overwritten (should still contain a formula or be None/from template)
        e2_val = ws["E2"].value
        # We just assert we didn't write a plain float/string to it
        check("E2 not overwritten with plain number (formula preserved or None)",
              not isinstance(e2_val, (int, float)) or e2_val == 0)

    except Exception as exc:
        print(f"  FAIL  Exception during populate_worksheet: {exc}")
        failed += 1

# --- Error handling ---
print("\n4. Error handling")
try:
    populate_worksheet(
        template_path=Path("/nonexistent/template.xlsx"),
        output_path=Path("/tmp/out.xlsx"),
        client_name="Test",
        year_end_date="April 30, 2026",
        tb_lines=[],
    )
    check("raises FileNotFoundError for missing template", False)
except FileNotFoundError:
    check("raises FileNotFoundError for missing template", True)
except Exception as e:
    check(f"raises FileNotFoundError (got {type(e).__name__})", False)

# --- Summary ---
print(f"\n{'-' * 40}")
print(f"  {passed + failed} checks: {passed} passed, {failed} failed")
if failed:
    sys.exit(1)
