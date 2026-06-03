"""
core/year_end_worksheet.py
Populate the year-end Excel worksheet template with a Sage 50 trial balance.

Template structure (R:\\Templates\\TEMPLATE_YearEnd_Accounting_Professional_BLANK_v2.xlsx):
  "0. Cover Sheet"  — D8=client, D9=year_end_date, D10=prepared_by, D11=date_prepared
  "1. Worksheet"    — rows 2-200 data, row 202 TOTALS, 203 Diff, 204 Balance Check
                      Cols A-D written by us; E-M are formula-driven (never touched)
  "2. Adjusting Entries", "3. Income Statement", "4. Balance Sheet" — untouched
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sage50.trial_balance_parser import TBLine

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_NUM_FMT   = "#,##0.00"
_NAVY      = "FF1F3864"   # header background (matches template)
_ROW_ALT   = "FFEAF2FB"   # alternating data row fill (light steel blue)
_ROW_WHITE = "FFFFFFFF"
_TOTALS_BG = "FFD9E1F2"   # totals row background

_WHITE_FONT  = Font(name="Calibri", bold=True,  color="FFFFFFFF", size=11)
_NORMAL_FONT = Font(name="Calibri", bold=False, color="FF000000", size=10)
_BOLD_FONT   = Font(name="Calibri", bold=True,  color="FF000000", size=10)
_ITALIC_FONT = Font(name="Calibri", bold=False, italic=True, color="FF404040", size=10)

_THIN  = Side(style="thin")
_MED   = Side(style="medium")

_THIN_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOTALS_BORDER = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

_COL_WIDTHS = {
    1: 10,   # A — Account No
    2: 38,   # B — Description
    3: 16,   # C — TB Debit
    4: 16,   # D — TB Credit
    5: 16,   # E — Adj Debit
    6: 16,   # F — Adj Credit
    7: 16,   # G — Adj TB Debit
    8: 16,   # H — Adj TB Credit
    9: 14,   # I — P&L Debit
    10: 14,  # J — P&L Credit
    11: 14,  # K — BS Debit
    12: 14,  # L — BS Credit
    13: 14,  # M — Type
}

# Template footer layout (original row numbers, zero-indexed from data block)
_ORIG_DATA_END   = 200   # last formula row in blank template
_ORIG_TOTALS_ROW = 202
_ORIG_DIFF_ROW   = 203
_ORIG_CHECK_ROW  = 204


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_header_row(ws) -> None:
    """Re-apply explicit white bold font to existing navy header (row 1)."""
    for col in range(1, 14):
        cell = ws.cell(row=1, column=col)
        cell.font      = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _style_data_row(ws, row: int, num_cols: int = 13) -> None:
    bg = _ROW_ALT if row % 2 == 0 else _ROW_WHITE
    fill = _fill(bg)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.font   = _NORMAL_FONT
        cell.border = _THIN_BORDER
        if col == 1:
            cell.alignment = Alignment(horizontal="center")
        elif col == 2:
            cell.alignment = Alignment(horizontal="left")
        elif col >= 3:
            cell.alignment = Alignment(horizontal="right")
        if col in (3, 4):
            cell.number_format = _NUM_FMT


def _style_footer_row(ws, row: int, label_col: int | None, bold: bool,
                      bg: str, num_cols: int = 12) -> None:
    fill = _fill(bg)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.border = _TOTALS_BORDER
        cell.font   = _BOLD_FONT if (bold or col == label_col) else _ITALIC_FONT
        if col >= 3:
            cell.alignment  = Alignment(horizontal="right")
            cell.number_format = _NUM_FMT
        elif col == 2:
            cell.alignment = Alignment(horizontal="left")


def _write_footer_formulas(ws, last_data_row: int) -> None:
    """Write TOTALS / Diff / Balance-Check rows with corrected row references."""
    totals_row = last_data_row + 3   # +1 skip, +2 spacers = 3 rows after last data
    diff_row   = totals_row + 1
    check_row  = totals_row + 2

    # TOTALS label
    ws.cell(row=totals_row, column=2).value = "TOTALS"

    # SUM formulas — cols C-L (3-12)
    for col in range(3, 13):
        ws.cell(row=totals_row, column=col).value = (
            f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last_data_row})"
        )

    # Diff row
    tr = totals_row
    ws.cell(row=diff_row, column=2).value = "Trial Balance Difference:"
    ws.cell(row=diff_row, column=3).value = f"=D{tr}-C{tr}"
    ws.cell(row=diff_row, column=4).value = f"=D{tr}-C{tr}"
    ws.cell(row=diff_row, column=5).value = (
        f'=IF(C{diff_row}=0,"BALANCED","Diff: "&TEXT(ABS(C{diff_row}),"#,##0.00"))'
    )
    ws.cell(row=diff_row, column=7).value  = f"=H{tr}-G{tr}"
    ws.cell(row=diff_row, column=8).value  = f"=H{tr}-G{tr}"
    ws.cell(row=diff_row, column=9).value  = f"=J{tr}-I{tr}"
    ws.cell(row=diff_row, column=12).value = f"=K{tr}-L{tr}"

    # Balance Check row
    dr = diff_row
    ws.cell(row=check_row, column=2).value  = "Adjusted Balance Check:"
    ws.cell(row=check_row, column=3).value  = f"=C{tr}+C{dr}"
    ws.cell(row=check_row, column=4).value  = f"=D{tr}"
    ws.cell(row=check_row, column=5).value  = (
        f'=IF(C{check_row}=D{check_row},"BALANCED","UNBALANCED")'
    )
    ws.cell(row=check_row, column=7).value  = f"=G{tr}+G{dr}"
    ws.cell(row=check_row, column=8).value  = f"=H{tr}"
    ws.cell(row=check_row, column=9).value  = f"=I{tr}+I{dr}"
    ws.cell(row=check_row, column=10).value = f"=J{tr}"
    ws.cell(row=check_row, column=11).value = f"=K{tr}"
    ws.cell(row=check_row, column=12).value = f"=L{tr}+L{dr}"

    return totals_row, diff_row, check_row


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def populate_worksheet(
    template_path: Path,
    output_path: Path,
    client_name: str,
    year_end_date: str,
    tb_lines: list[TBLine],
    prepared_by: str = "Jorge Quinonez CPA",
) -> Path:
    """Populate the year-end Excel template with trial balance data.

    Writes Cover Sheet metadata (D8-D11) and Worksheet tab TB columns (A-D).
    Deletes unused template formula rows so the TOTALS section immediately
    follows the data (with 2 spacer rows). Applies number formatting and
    professional styling throughout.

    Returns output_path for caller convenience.
    Raises FileNotFoundError / ValueError on missing template or tabs.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Year-end template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)

    # --- Cover Sheet ---
    cover_tab = "0. Cover Sheet"
    if cover_tab not in wb.sheetnames:
        raise ValueError(
            f"Template is missing tab '{cover_tab}'. Found: {wb.sheetnames}"
        )
    cover = wb[cover_tab]
    cover["D8"] = client_name
    cover["D9"] = year_end_date
    cover["D9"].number_format = "General"   # template cell is mm-dd-yy; force text
    cover["D10"] = prepared_by
    cover["D11"] = date.today().strftime("%B %d, %Y")

    # --- Worksheet tab ---
    ws_tab = "1. Worksheet"
    if ws_tab not in wb.sheetnames:
        raise ValueError(
            f"Template is missing tab '{ws_tab}'. Found: {wb.sheetnames}"
        )
    ws = wb[ws_tab]

    # 1. Write TB data to rows 2..n+1
    last_data_row = len(tb_lines) + 1   # row index of last written data row
    for i, line in enumerate(tb_lines):
        row = i + 2
        ws.cell(row=row, column=1).value = line.account_no
        ws.cell(row=row, column=2).value = line.description
        ws.cell(row=row, column=3).value = float(line.debit)  if line.debit  else None
        ws.cell(row=row, column=4).value = float(line.credit) if line.credit else None
        # Cols E-M (5-13) are pre-built SUMIF/calc formulas — never touch

    # 2. Delete the unused formula rows between last data row + 2 spacers and
    #    the original TOTALS row (202), then also the original blank spacer (201).
    #    We keep last_data_row+2 and last_data_row+3 as blank spacers.
    delete_start = last_data_row + 4           # first row to remove
    delete_end   = _ORIG_TOTALS_ROW - 1        # original row just before TOTALS (= 201)
    rows_to_delete = delete_end - delete_start + 1
    if rows_to_delete > 0:
        ws.delete_rows(delete_start, rows_to_delete)

    # 3. Rewrite footer formulas (references are now at new row positions).
    #    delete_rows shifts rows 202-204 up by rows_to_delete.
    totals_row, diff_row, check_row = _write_footer_formulas(ws, last_data_row)

    # 4. Apply column widths
    _apply_col_widths(ws)

    # 5. Style header row (keep navy fill, add explicit white font)
    _style_header_row(ws)

    # 6. Style data rows
    for row in range(2, last_data_row + 1):
        _style_data_row(ws, row)

    # 7. Style footer rows
    _style_footer_row(ws, totals_row, label_col=2, bold=True,  bg=_TOTALS_BG)
    _style_footer_row(ws, diff_row,   label_col=2, bold=False, bg="FFEFF4FB")
    _style_footer_row(ws, check_row,  label_col=2, bold=False, bg="FFEFF4FB")

    # 8. Freeze top row
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
